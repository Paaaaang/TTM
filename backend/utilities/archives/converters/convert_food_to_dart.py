#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
음식 데이터를 Dart 코드로 변환
"""
import openpyxl
import re
import os

excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "nutrition_db.xlsx")

wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

# 음식 분류 규칙 (음식명 기반으로 카테고리 자동 분류)
def categorize_food(food_name):
    """음식명을 보고 카테고리 자동 분류"""
    if any(x in food_name for x in ['밥', '죽', '김밥', '비빔밥', '덮밥', '볶음밥']):
        return '밥류'
    elif any(x in food_name for x in ['국', '탕', '찌개', '전골', '육수']):
        return '국/찌개'
    elif any(x in food_name for x in ['고기', '삼겹살', '갈비', '불고기', '닭', '오리', '양념치킨', '후라이드', '치킨']):
        return '고기/생선'
    elif any(x in food_name for x in ['생선', '고등어', '갈치', '삼치', '참치', '연어', '광어', '가자미', '조기', '명태', '대구']):
        return '고기/생선'
    elif any(x in food_name for x in ['계란', '달걀', '계란찜', '에그', '오믈렛']):
        return '고기/생선'
    elif any(x in food_name for x in ['샐러드', '나물', '김치', '무침', '겉절이', '쌈', '상추', '시금치', '콩나물', '숙주', '미나리']):
        return '채소/샐러드'
    elif any(x in food_name for x in ['빵', '식빵', '크루아상', '케이크', '파이', '쿠키', '도넛', '머핀', '타르트']):
        return '빵/디저트'
    elif any(x in food_name for x in ['초콜릿', '아이스크림', '푸딩', '젤리', '캔디', '사탕', '과자']):
        return '빵/디저트'
    elif any(x in food_name for x in ['커피', '라떼', '아메리카노', '카푸치노', '음료', '주스', '차', '티', '콜라', '사이다', '우유', '두유']):
        return '음료'
    elif any(x in food_name for x in ['과일', '사과', '바나나', '오렌지', '포도', '딸기', '수박', '메론', '키위', '망고']):
        return '간식'
    elif any(x in food_name for x in ['요거트', '요구르트', '치즈', '견과류', '땅콩', '아몬드', '호두']):
        return '간식'
    elif any(x in food_name for x in ['면', '라면', '우동', '짜장', '짬뽕', '파스타', '스파게티', '국수', '냉면', '칼국수']):
        return '면류'
    else:
        return '기타'

print("Dart 코드 생성 중...")
print("=" * 80)

dart_code = []
dart_code.append("  // 음식 데이터베이스 (총 {}개)".format(sheet.max_row - 1))
dart_code.append("  final List<FoodItem> _foodDatabase = [")

row_count = 0
for row in sheet.iter_rows(min_row=2, values_only=True):
    food_name = row[0]  # 음식명
    weight = row[1]     # 중량(g)
    calories = row[2]   # 에너지(kcal)
    carbs = row[3]      # 탄수화물(g)
    fat = row[5]        # 지방(g)
    protein = row[6]    # 단백질(g)
    
    if not food_name or not weight or not calories:
        continue
    
    # 값 정리
    try:
        weight = float(weight) if weight else 100.0
        calories = int(float(calories)) if calories else 0
        carbs = float(carbs) if carbs and carbs != '-' else 0.0
        fat = float(fat) if fat and fat != '-' else 0.0
        protein = float(protein) if protein and protein != '-' else 0.0
    except (ValueError, TypeError):
        continue
    
    # 카테고리 자동 분류
    category = categorize_food(food_name)
    
    # Dart 코드 생성
    dart_line = "    FoodItem(name: '{}', calories: {}, category: '{}', grams: {:.0f}, carbs: {:.1f}, protein: {:.1f}, fat: {:.1f}),".format(
        food_name, calories, category, weight, carbs, protein, fat
    )
    dart_code.append(dart_line)
    row_count += 1

dart_code.append("  ];")

# 출력
output_file = r'C:\Users\smhrd\Desktop\App\ttm\backend\food_database_dart.txt'
with open(output_file, 'w', encoding='utf-8') as f:
    f.write('\n'.join(dart_code))

print(f"✅ 총 {row_count}개의 음식 데이터 변환 완료")
print(f"📁 파일 저장: {output_file}")

# 카테고리별 통계
from collections import Counter
categories = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        categories.append(categorize_food(row[0]))

cat_stats = Counter(categories)
print(f"\n📊 카테고리별 통계:")
for cat, count in sorted(cat_stats.items(), key=lambda x: x[1], reverse=True):
    print(f"  {cat}: {count}개")

print("=" * 80)
wb.close()
