#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
음식 데이터 엑셀 파일 읽기
"""
import openpyxl
import json
import os

excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "nutrition_db.xlsx")

wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

print("=" * 80)
print("음식 데이터 엑셀 파일 구조")
print("=" * 80)

# 헤더 읽기
headers = []
for cell in sheet[1]:
    headers.append(cell.value)
print(f"\n📋 컬럼명 ({len(headers)}개):")
for i, h in enumerate(headers, 1):
    print(f"  {i}. {h}")

# 첫 10행 데이터 샘플
print(f"\n📊 데이터 샘플 (처음 10행):")
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=11, values_only=True), 2):
    print(f"\nRow {row_idx}:")
    for col_idx, (header, value) in enumerate(zip(headers, row)):
        if value:
            print(f"  {header}: {value}")

# 카테고리 목록 추출
categories = set()
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:  # 첫 번째 컬럼이 카테고리라고 가정
        categories.add(row[0])

print(f"\n📂 발견된 카테고리 ({len(categories)}개):")
for cat in sorted(categories)[:20]:
    print(f"  - {cat}")

print(f"\n📈 총 데이터 행 수: {sheet.max_row - 1}")
print("=" * 80)

wb.close()
